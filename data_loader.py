import os
import glob
import openpyxl
import pandas as pd

BASE_PATH = os.path.join(os.path.dirname(__file__), "data")

REMICON_FACTORIES = [
    '안양', '인천', '파주', '김포', '부산', '서부산', '김해',
    '정관', '양산', '창원', '대구', '울산', '아산', '전주',
    '군산', '원주', '제주'
]

REGION_ROWS = ['수도권', '영남권', '중부권', '레미콘 계']
SUMMARY_ROWS = ['건자재', '골재 계', '기타', '합계']


def get_available_years():
    years = []
    for name in os.listdir(BASE_PATH):
        full = os.path.join(BASE_PATH, name)
        if os.path.isdir(full) and name.isdigit() and len(name) == 4:
            years.append(name)
    return sorted(years, reverse=True)


def get_available_months(year):
    months = []
    year_folder = os.path.join(BASE_PATH, str(year))
    if not os.path.exists(year_folder):
        return months
    for f in os.listdir(year_folder):
        if f.startswith('손익보고서(') and f.endswith('.xlsx') and not f.startswith('~$'):
            try:
                month = f.split('-')[1].replace(').xlsx', '')
                months.append(int(month))
            except:
                pass
    return sorted(set(months))


def find_report_file(year, month):
    month_str = f"{int(month):02d}"
    path = os.path.join(BASE_PATH, str(year), f"손익보고서({year}-{month_str}).xlsx")
    return path if os.path.exists(path) else None


def load_factory_data(year, month, period="당월"):
    filepath = find_report_file(year, month)
    if not filepath:
        return None

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    sheet_name = f'사업장별({period})'
    rows_data = []
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=7, max_row=50, values_only=True):
            name = row[8] if (row[8] and isinstance(row[8], str)) else (row[7] if (len(row) > 7 and row[7] and isinstance(row[7], str)) else None)
            if name and isinstance(name, str):
                rows_data.append({
                    '공장명': name,
                    '물량_계획': row[9],
                    '물량_실적': row[10],
                    '물량_차이': row[11],
                    '물량_전년': row[12],
                    '매출_계획': row[14],
                    '매출_실적': row[15],
                    '매출_차이': row[16],
                    '매출_전년': row[17],
                    '영업이익_계획': row[19],
                    '영업이익_실적': row[20],
                    '영업이익_차이': row[21],
                    '영업이익_전년': row[22],
                    '판매단가_계획': row[24],
                    '판매단가_실적': row[25],
                    '판매단가_전년': row[26],
                    '변동비_계획': row[27],
                    '변동비_실적': row[28],
                    '변동비_전년': row[29],
                    '공헌이익_계획': row[30],
                    '공헌이익_실적': row[31],
                    '공헌이익_전년': row[32],
                })
    wb.close()

    df = pd.DataFrame(rows_data)
    # 합계 행이 2개(레미콘만/전체)인 경우 마지막(건자재 포함 전체) 행만 유지
    df = df.drop_duplicates(subset=['공장명'], keep='last')
    df['연도'] = int(year)
    df['월'] = int(month)
    return df


def load_remicon_factories(year, month):
    df = load_factory_data(year, month)
    if df is None:
        return None
    return df[df['공장명'].isin(REMICON_FACTORIES)].copy()


def load_summary_data(year, month):
    df = load_factory_data(year, month)
    if df is None:
        return None
    summary_names = ['레미콘 계', '건자재', '골재 계', '기타', '합계']
    return df[df['공장명'].isin(summary_names)].copy()


def load_sijangbyul_raw(year, month, period="당월"):
    """사업장별 시트에서 모든 데이터 행을 순서대로 반환 (중복 제거 없이)."""
    filepath = find_report_file(year, month)
    if not filepath:
        return []

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_name = f'사업장별({period})'
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[sheet_name]
    rows_data = []
    prev_section = ''

    # 섹션 식별용 앞쪽 컬럼 범위
    SECTION_MAP = {
        '레미콘': '레미콘', '직영': '레미콘', '골재': '골재',
        '기타': '기타', '건자재': '건자재',
    }

    for row in ws.iter_rows(min_row=7, max_row=50, values_only=True):
        if len(row) < 20:
            continue

        # 이름 찾기 (col I=8, col H=7, col J=9 순서)
        name = None
        for ci in [8, 7, 9]:
            v = row[ci] if ci < len(row) else None
            if v and isinstance(v, str) and v.strip():
                name = v.strip()
                break
        if not name:
            continue

        # 섹션 레이블 (앞쪽 컬럼들에서 찾기)
        section = prev_section
        for ci in range(0, 7):
            v = row[ci] if ci < len(row) else None
            if v and isinstance(v, str):
                s = v.strip()
                if s in SECTION_MAP:
                    section = SECTION_MAP[s]
                    break
        prev_section = section

        def n(v):
            if v is None:
                return None
            try:
                return float(v)
            except Exception:
                return None

        rows_data.append({
            '구분': name,
            '섹션': section,
            '물량_계획':        n(row[9]),
            '물량_실적':        n(row[10]),
            '물량_차이':        n(row[11]),
            '물량_전년':        n(row[12]),
            '물량_전년차이':    n(row[13]),
            '매출_계획':        n(row[14]),
            '매출_실적':        n(row[15]),
            '매출_차이':        n(row[16]),
            '매출_전년':        n(row[17]),
            '매출_전년차이':    n(row[18]),
            '영업이익_계획':    n(row[19]),
            '영업이익_실적':    n(row[20]),
            '영업이익_차이':    n(row[21]),
            '영업이익_전년':    n(row[22]),
            '영업이익_전년차이':n(row[23]),
            '판매단가_계획':    n(row[24]),
            '판매단가_실적':    n(row[25]),
            '판매단가_전년':    n(row[26]),
            '변동비_계획':      n(row[27]),
            '변동비_실적':      n(row[28]),
            '변동비_전년':      n(row[29]),
            '공헌이익_계획':    n(row[30]),
            '공헌이익_실적':    n(row[31]),
            '공헌이익_전년':    n(row[32]) if len(row) > 32 else None,
        })

    wb.close()
    return rows_data


def load_overview(year, month):
    """손익총괄 시트에서 레미콘/건자재/골재/기타/합계 당월+누계 데이터를 읽어 반환."""
    filepath = find_report_file(year, month)
    if not filepath:
        return None

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if '손익총괄' not in wb.sheetnames:
        wb.close()
        return None

    ws = wb['손익총괄']
    rows = {i: row for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1)}
    wb.close()

    def v(row_idx, col_idx):
        r = rows.get(row_idx)
        if r is None or col_idx >= len(r):
            return None
        val = r[col_idx]
        return float(val) if val is not None else None

    # 컬럼 인덱스: [4]=당월계획 [5]=당월실적 [6]=당월차이 [7]=당월전년
    #              [9]=누계계획 [10]=누계실적 [11]=누계차이 [12]=누계전년
    items = []
    for name, r_vol, r_sale, r_oi in [
        ('레미콘', 7,  9, 25),
        ('골재',   8, 10, 27),
        ('건자재', None, 33, 36),
        ('기타',   None, 11, 28),
        ('합계',   None, 12, 29),
    ]:
        row = {
            '구분': name,
            '물량_계획':   v(r_vol, 4) if r_vol else None,
            '물량_실적':   v(r_vol, 5) if r_vol else None,
            '물량_차이':   v(r_vol, 6) if r_vol else None,
            '물량_전년':   v(r_vol, 7) if r_vol else None,
            '물량_누계계획': v(r_vol, 9) if r_vol else None,
            '물량_누계실적': v(r_vol, 10) if r_vol else None,
            '물량_누계차이': v(r_vol, 11) if r_vol else None,
            '물량_누계전년': v(r_vol, 12) if r_vol else None,
            '매출_계획':   v(r_sale, 4),
            '매출_실적':   v(r_sale, 5),
            '매출_차이':   v(r_sale, 6),
            '매출_전년':   v(r_sale, 7),
            '매출_누계계획': v(r_sale, 9),
            '매출_누계실적': v(r_sale, 10),
            '매출_누계차이': v(r_sale, 11),
            '매출_누계전년': v(r_sale, 12),
            '영업이익_계획':   v(r_oi, 4),
            '영업이익_실적':   v(r_oi, 5),
            '영업이익_차이':   v(r_oi, 6),
            '영업이익_전년':   v(r_oi, 7),
            '영업이익_누계계획': v(r_oi, 9),
            '영업이익_누계실적': v(r_oi, 10),
            '영업이익_누계차이': v(r_oi, 11),
            '영업이익_누계전년': v(r_oi, 12),
        }
        items.append(row)

    df = pd.DataFrame(items)

    # 합계 행에 건자재 합산 (엑셀 합계 행이 건자재를 포함하지 않음)
    jc_row = df[df['구분'] == '건자재']
    tot_idx = df.index[df['구분'] == '합계'].tolist()
    if not jc_row.empty and tot_idx:
        ti = tot_idx[0]
        for col in ['매출_계획','매출_실적','매출_차이','매출_전년',
                    '매출_누계계획','매출_누계실적','매출_누계차이','매출_누계전년',
                    '영업이익_계획','영업이익_실적','영업이익_차이','영업이익_전년',
                    '영업이익_누계계획','영업이익_누계실적','영업이익_누계차이','영업이익_누계전년']:
            jv = jc_row.iloc[0].get(col)
            tv = df.at[ti, col]
            if jv is not None and tv is not None:
                df.at[ti, col] = tv + jv
            elif jv is not None:
                df.at[ti, col] = jv

    return df
